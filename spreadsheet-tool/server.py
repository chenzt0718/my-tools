"""Spreadsheet Tool Backend - Flask server with file I/O, validation, and statistics."""

import io
import json
import os
import re
from pathlib import Path

import tempfile

import openpyxl
import pandas as pd
from flask import Flask, jsonify, render_template, request

from crypto_stub import (
    decrypt_file,
    delete_temp_file,
    encrypt_file,
    fetch_decrypt_key,
    fetch_encrypt_key,
    is_encrypted_file,
    strip_encryption_suffix,
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB max upload

# In-memory store for the currently open workbook
_current_path: str | None = None
_current_wb: openpyxl.Workbook | None = None
_modified_cells: dict = {}  # key: "SheetName!col,row" -> new_value
_cell_colors: dict = {}     # key: "SheetName!col,row" -> color hex
_encryption_info: dict | None = None  # 加密上下文: {encrypted, original_filename, key}


def col_letter(idx: int) -> str:
    """0-indexed column number to Excel column letter."""
    result = ""
    while idx >= 0:
        result = chr(65 + (idx % 26)) + result
        idx = idx // 26 - 1
    return result


def col_index(letter: str) -> int:
    """Excel column letter to 0-indexed column number."""
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - 64)
    return result - 1


def read_sheet_data(ws, sheet_name: str) -> dict:
    """Read all data from a worksheet into a JSON-serializable structure."""
    rows = []
    max_col = ws.max_column or 1
    max_row = ws.max_row or 1

    # Read merged cells info
    merged = []
    for m in ws.merged_cells.ranges:
        merged.append({
            "min_row": m.min_row, "max_row": m.max_row,
            "min_col": m.min_col, "max_col": m.max_col,
        })

    col_widths = {}
    for col_letter_key, dim in ws.column_dimensions.items():
        if dim.width:
            col_widths[col_letter_key] = dim.width

    row_heights = {}
    for row_num, dim in ws.row_dimensions.items():
        if dim.height:
            row_heights[row_num] = dim.height

    for row_idx in range(1, max_row + 1):
        cells = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            bg_color = None
            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                rgb = str(cell.fill.fgColor.rgb)
                if len(rgb) == 8 and rgb != "00000000":
                    bg_color = "#" + rgb[2:]

            font_bold = cell.font.bold if cell.font else False

            # Check for modifications
            cell_key = f"{sheet_name}!{col_idx},{row_idx}"
            if cell_key in _modified_cells:
                value = _modified_cells[cell_key]
            if cell_key in _cell_colors:
                bg_color = _cell_colors[cell_key]

            cells.append({
                "value": value,
                "bg": bg_color,
                "bold": font_bold,
            })
        rows.append(cells)

    return {
        "name": sheet_name,
        "rows": rows,
        "max_row": max_row,
        "max_col": max_col,
        "merged": merged,
        "col_widths": col_widths,
        "row_heights": row_heights,
    }


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/open", methods=["POST"])
def open_file():
    global _current_wb, _current_path, _modified_cells, _cell_colors
    data = request.get_json()
    filepath = data.get("path", "")

    if not filepath or not os.path.exists(filepath):
        return jsonify({"error": f"文件不存在: {filepath}"}), 400

    try:
        _current_wb = openpyxl.load_workbook(filepath)
        _current_path = filepath
        _modified_cells = {}
        _cell_colors = {}

        sheets = []
        for name in _current_wb.sheetnames:
            ws = _current_wb[name]
            sheets.append(read_sheet_data(ws, name))

        return jsonify({
            "filename": os.path.basename(filepath),
            "sheets": sheets,
            "active_sheet": _current_wb.active.title if _current_wb.active else sheets[0]["name"],
        })
    except Exception as e:
        return jsonify({"error": f"无法打开文件: {str(e)}"}), 500


@app.route("/api/open_file", methods=["POST"])
def open_uploaded_file():
    """Handle file upload — supports both plain and encrypted files."""
    global _current_wb, _current_path, _modified_cells, _cell_colors, _encryption_info

    if "file" not in request.files:
        return jsonify({"error": "未提供文件"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "文件名为空"}), 400

    # Clean up previous encrypted session if any
    if _encryption_info and _encryption_info.get("encrypted") and _current_path:
        delete_temp_file(_current_path)

    tmp_path = None
    decrypted_path = None
    encrypted = False

    try:
        # Step 1: Save uploaded file to temp location
        suffix = os.path.splitext(file.filename)[1] or ".xlsx"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        file.save(tmp.name)
        tmp.close()
        tmp_path = tmp.name

        # Step 2: Check if file is encrypted
        encrypted = is_encrypted_file(file.filename)

        if encrypted:
            # Step 2a: Fetch decryption key from API
            decrypt_key = fetch_decrypt_key(file.filename)

            # Step 2b: Decrypt to temporary file
            decrypted_path = decrypt_file(tmp_path, decrypt_key)

            # Step 2c: Open the decrypted file
            _current_wb = openpyxl.load_workbook(decrypted_path)
            _current_path = decrypted_path  # 指向解密后的临时文件（保存时需要）

            # Step 2d: Record encryption context for save
            _encryption_info = {
                "encrypted": True,
                "original_filename": strip_encryption_suffix(file.filename),
                "uploaded_filename": file.filename,
                "key": decrypt_key,
                "uploaded_tmp": tmp_path,  # 保存时覆盖此文件
            }
        else:
            # Plain file — open directly
            _current_wb = openpyxl.load_workbook(tmp_path)
            _current_path = tmp_path
            _encryption_info = None

        _modified_cells = {}
        _cell_colors = {}

        sheets = []
        for name in _current_wb.sheetnames:
            ws = _current_wb[name]
            sheets.append(read_sheet_data(ws, name))

        return jsonify({
            "filename": file.filename,
            "display_name": _encryption_info["original_filename"] if _encryption_info else file.filename,
            "sheets": sheets,
            "active_sheet": _current_wb.active.title if _current_wb.active else sheets[0]["name"],
            "encrypted": encrypted,
        })
    except Exception as e:
        return jsonify({"error": f"无法打开文件: {str(e)}"}), 500
    finally:
        # Step 3: Clean up — delete the uploaded encrypted temp file if decrypted
        # Keep decrypted_path (it's _current_path now, will be cleaned on save/shutdown)
        if encrypted and decrypted_path and tmp_path and tmp_path != decrypted_path:
            delete_temp_file(tmp_path)


@app.route("/api/save", methods=["POST"])
def save_file():
    global _current_wb, _current_path, _modified_cells, _encryption_info
    if _current_wb is None:
        return jsonify({"error": "没有打开的文件"}), 400

    tmp_save_path = None
    encrypted_path = None

    try:
        # Step 1: Apply modifications to workbook
        for cell_key, new_value in _modified_cells.items():
            sheet_name, coords = cell_key.split("!")
            col_str, row_str = coords.split(",")
            col_idx = int(col_str)
            row_idx = int(row_str)
            ws = _current_wb[sheet_name]
            ws.cell(row=row_idx, column=col_idx).value = new_value

        # Apply cell colors
        for cell_key, color in _cell_colors.items():
            sheet_name, coords = cell_key.split("!")
            col_str, row_str = coords.split(",")
            col_idx = int(col_str)
            row_idx = int(row_str)
            ws = _current_wb[sheet_name]
            from openpyxl.styles import PatternFill
            if color:
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=color.lstrip("#"), end_color=color.lstrip("#"),
                    fill_type="solid"
                )
            else:
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(fill_type=None)

        # Step 2: Save workbook to a temp clean file
        tmp_save = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_save.close()
        tmp_save_path = tmp_save.name
        _current_wb.save(tmp_save_path)

        if _encryption_info and _encryption_info.get("encrypted"):
            # ── Encrypted save flow ──────────────────────────────
            # Step 3a: Fetch encrypt key from API
            encrypt_key = fetch_encrypt_key(_encryption_info["original_filename"])

            # Step 3b: Encrypt the saved file
            encrypted_path = encrypt_file(tmp_save_path, encrypt_key)

            # Step 3c: Replace the original encrypted file
            target = _encryption_info.get("uploaded_tmp")
            if target and os.path.exists(target):
                os.unlink(target)
            # In real scenario, the encrypted file would be uploaded back
            # via API or saved to original location. Here we just copy.
            import shutil
            shutil.copy(encrypted_path, target)

            display_name = _encryption_info["original_filename"]
            result_msg = "加密保存成功"
        else:
            # ── Plain save flow ──────────────────────────────────
            # Step 3: If current_path is a temp file, also update it
            import shutil
            shutil.copy(tmp_save_path, _current_path)
            display_name = os.path.basename(_current_path)
            result_msg = "保存成功"

        _modified_cells = {}
        return jsonify({
            "message": result_msg,
            "filename": display_name,
            "encrypted": bool(_encryption_info and _encryption_info.get("encrypted")),
        })
    except Exception as e:
        return jsonify({"error": f"保存失败: {str(e)}"}), 500
    finally:
        # Clean up intermediate temp files only (not the working decrypted file)
        if tmp_save_path:
            delete_temp_file(tmp_save_path)
        if encrypted_path:
            delete_temp_file(encrypted_path)


@app.route("/api/cell/update", methods=["POST"])
def update_cell():
    data = request.get_json()
    sheet = data["sheet"]
    col = data["col"]  # 1-indexed
    row = data["row"]  # 1-indexed
    value = data["value"]

    cell_key = f"{sheet}!{col},{row}"
    _modified_cells[cell_key] = value
    return jsonify({"status": "ok"})


@app.route("/api/cell/color", methods=["POST"])
def set_cell_color():
    data = request.get_json()
    sheet = data["sheet"]
    col = data["col"]
    row = data["row"]
    color = data.get("color")  # hex string or null to clear

    cell_key = f"{sheet}!{col},{row}"
    if color:
        _cell_colors[cell_key] = color
    else:
        _cell_colors.pop(cell_key, None)
    return jsonify({"status": "ok"})


@app.route("/api/validate", methods=["POST"])
def validate_data():
    """Run validation rules on the current sheet data and return issues."""
    data = request.get_json()
    rules = data.get("rules", [])
    sheet_data = data.get("sheet_data", [])

    issues = []

    for rule in rules:
        rule_type = rule.get("type")
        target = rule.get("target", "all")  # "all", "column:N", "row:N", "range"

        for row_idx, row_cells in enumerate(sheet_data):
            for col_idx, cell in enumerate(row_cells):
                if not _match_target(target, row_idx + 1, col_idx + 1):
                    continue

                val = cell.get("value")

                if rule_type == "not_empty":
                    if val is None or str(val).strip() == "":
                        issues.append({
                            "row": row_idx + 1, "col": col_idx + 1,
                            "message": f"单元格 {col_letter(col_idx)}{row_idx + 1} 不能为空",
                            "rule": "not_empty",
                        })

                elif rule_type == "is_number":
                    if val is not None and str(val).strip() != "":
                        try:
                            float(val)
                        except (ValueError, TypeError):
                            issues.append({
                                "row": row_idx + 1, "col": col_idx + 1,
                                "message": f"单元格 {col_letter(col_idx)}{row_idx + 1} 应为数值",
                                "rule": "is_number",
                            })

                elif rule_type == "range":
                    min_val = rule.get("min")
                    max_val = rule.get("max")
                    if val is not None and str(val).strip() != "":
                        try:
                            num = float(val)
                            if min_val is not None and num < min_val:
                                issues.append({
                                    "row": row_idx + 1, "col": col_idx + 1,
                                    "message": f"单元格 {col_letter(col_idx)}{row_idx + 1} 值 {num} 小于最小值 {min_val}",
                                    "rule": "range",
                                })
                            if max_val is not None and num > max_val:
                                issues.append({
                                    "row": row_idx + 1, "col": col_idx + 1,
                                    "message": f"单元格 {col_letter(col_idx)}{row_idx + 1} 值 {num} 大于最大值 {max_val}",
                                    "rule": "range",
                                })
                        except (ValueError, TypeError):
                            pass

                elif rule_type == "regex":
                    pattern = rule.get("pattern", "")
                    if val is not None and str(val).strip() != "" and pattern:
                        if not re.match(pattern, str(val)):
                            issues.append({
                                "row": row_idx + 1, "col": col_idx + 1,
                                "message": f"单元格 {col_letter(col_idx)}{row_idx + 1} 不匹配格式 {pattern}",
                                "rule": "regex",
                            })

    return jsonify({"issues": issues, "count": len(issues)})


def _match_target(target: str, row: int, col: int) -> bool:
    if target == "all":
        return True
    if target.startswith("column:"):
        t_col = int(target.split(":")[1])
        return col == t_col
    if target.startswith("row:"):
        t_row = int(target.split(":")[1])
        return row == t_row
    return True


@app.route("/api/stats", methods=["POST"])
def compute_stats():
    """Compute statistics for selected cells."""
    data = request.get_json()
    cells = data.get("cells", [])  # list of {row, col, value}

    values = []
    for c in cells:
        v = c.get("value")
        if v is not None and str(v).strip() != "":
            try:
                values.append(float(v))
            except (ValueError, TypeError):
                pass

    if not values:
        return jsonify({"count": 0, "message": "选中区域无有效数值"})

    import statistics
    return jsonify({
        "count": len(values),
        "sum": round(sum(values), 4),
        "avg": round(statistics.mean(values), 4),
        "median": round(statistics.median(values), 4),
        "min": round(min(values), 4),
        "max": round(max(values), 4),
        "stdev": round(statistics.stdev(values), 4) if len(values) > 1 else 0,
    })


@app.route("/api/encryption_status", methods=["GET"])
def encryption_status():
    """Return current encryption state for the frontend."""
    if _encryption_info and _encryption_info.get("encrypted"):
        return jsonify({
            "encrypted": True,
            "original_filename": _encryption_info.get("original_filename", ""),
        })
    return jsonify({"encrypted": False})


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
